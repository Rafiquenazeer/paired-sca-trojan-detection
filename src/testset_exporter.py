"""
testset_exporter.py  (v6 update)
===================================
Exports generated MERS-family test vectors in human-readable formats.

For each run, files are written under:

    <vectors_dir>/<circuit_name>/

Generated files
---------------
  * <circuit>_<method>_vectors.txt
      Binary vector strings, one vector per line, no header.
      The bit order is exactly circuit.primary_inputs.

  * <circuit>_<method>_vectors_hex.txt
      Hex version of the same vectors, one vector per line, no header.

  * <circuit>_<method>_eval10k_vectors.txt / _hex.txt
      Only written when the full testset has more than 10,000 vectors.
      These are the first 10,000 vectors used for fair comparison in the paper.

  * <circuit>_all_testsets.xlsx
      Compact Excel workbook with one sheet per method.  Each row contains:
      vector index, binary bitstring, hex string, and vector width.

  * primary_input_order.txt
      Maps bit index to primary-input / scan-input name.

  * vector_export_manifest.json
      Machine-readable summary of exported files and vector counts.

Important
---------
Do not shuffle vectors before side-channel/power experiments.  For MERS,
MERS-h, and MERS-s, the order of vectors matters because switching is created
by transitions between consecutive vectors.
"""

from __future__ import annotations

import json
import logging
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, Iterable, List, Sequence, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

MAX_EVAL_VECTORS = 10_000


class TestsetExporter:
    """Export test vectors to .txt and .xlsx formats."""

    def __init__(self, output_dir: str = "mers_testvectors"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(self,
                 testsets: Dict[str, Sequence[Sequence[int]]],
                 circuit,
                 circuit_name: str,
                 eval_limit: int = MAX_EVAL_VECTORS) -> dict:
        """
        Export all provided testsets.

        Parameters
        ----------
        testsets:
            Mapping such as {"Random": random_testset, "MERS": mers_testset, ...}
        circuit:
            Parsed Circuit object.  Must provide primary_inputs.
        circuit_name:
            Used in file and folder names.
        eval_limit:
            First-N vectors used by the paper for fair comparison when a testset
            is larger than 10K.

        Returns
        -------
        dict
            Manifest with output paths and vector counts.
        """
        run_dir = self.output_dir / _safe_filename(circuit_name)
        run_dir.mkdir(parents=True, exist_ok=True)

        input_names = list(getattr(circuit, "primary_inputs", []))
        width = len(input_names)

        logger.info("[VectorExport] Exporting test vectors to %s", run_dir)

        manifest = {
            "circuit_name": circuit_name,
            "exported_at": datetime.now().isoformat(timespec="seconds"),
            "vector_width": width,
            "bit_order": "left-to-right matches circuit.primary_inputs; bit 0 is the first name in primary_input_order.txt",
            "important_note": "Do not shuffle vectors. MERS/MERS-h/MERS-s depend on consecutive-vector transitions.",
            "output_directory": str(run_dir),
            "primary_input_order_file": None,
            "excel_workbook": None,
            "methods": {},
        }

        # Save input order first.  This is essential for future use.
        input_order_path = run_dir / "primary_input_order.txt"
        with open(input_order_path, "w", encoding="utf-8") as fh:
            fh.write("# Bit index -> primary input / scan input name\n")
            fh.write("# Vector bitstrings use this exact order.\n")
            for idx, name in enumerate(input_names):
                fh.write(f"{idx}\t{name}\n")
        manifest["primary_input_order_file"] = str(input_order_path)

        readme_path = run_dir / "README_VECTOR_EXPORTS.txt"
        self._write_readme(readme_path, circuit_name, width)

        # Create compact Excel workbook.
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"
        _setup_summary_sheet(ws_summary)

        ws_order = wb.create_sheet("Input Order")
        _setup_input_order_sheet(ws_order, input_names)

        summary_rows = []
        for method, raw_testset in testsets.items():
            if raw_testset is None:
                continue

            testset = _normalise_testset(raw_testset)
            if not testset:
                continue

            method_safe = _safe_filename(method)
            method_sheet = _safe_sheet_name(method)

            bin_path = run_dir / f"{_safe_filename(circuit_name)}_{method_safe}_vectors.txt"
            hex_path = run_dir / f"{_safe_filename(circuit_name)}_{method_safe}_vectors_hex.txt"
            self._write_binary_txt(bin_path, testset)
            self._write_hex_txt(hex_path, testset)

            method_info = {
                "vector_count_full": len(testset),
                "vector_width": len(testset[0]) if testset else width,
                "binary_txt": str(bin_path),
                "hex_txt": str(hex_path),
                "excel_sheet": method_sheet,
                "eval10k_binary_txt": None,
                "eval10k_hex_txt": None,
                "eval10k_count": min(len(testset), eval_limit),
            }

            # Export the first 10K separately only when the full testset is larger.
            # These are the vectors used in the paper-style evaluation.
            if len(testset) > eval_limit:
                eval_ts = testset[:eval_limit]
                eval_bin = run_dir / f"{_safe_filename(circuit_name)}_{method_safe}_eval10k_vectors.txt"
                eval_hex = run_dir / f"{_safe_filename(circuit_name)}_{method_safe}_eval10k_vectors_hex.txt"
                self._write_binary_txt(eval_bin, eval_ts)
                self._write_hex_txt(eval_hex, eval_ts)
                method_info["eval10k_binary_txt"] = str(eval_bin)
                method_info["eval10k_hex_txt"] = str(eval_hex)

            # Add one compact sheet per method.
            ws = wb.create_sheet(method_sheet)
            _setup_vector_sheet(ws, method, testset)

            manifest["methods"][method] = method_info
            summary_rows.append([
                method,
                len(testset),
                method_info["eval10k_count"],
                method_info["vector_width"],
                str(bin_path.name),
                str(hex_path.name),
                method_sheet,
                "Yes" if len(testset) > eval_limit else "No",
            ])

        _fill_summary_sheet(ws_summary, circuit_name, width, summary_rows, readme_path.name)

        excel_path = run_dir / f"{_safe_filename(circuit_name)}_all_testsets.xlsx"
        wb.save(excel_path)
        manifest["excel_workbook"] = str(excel_path)
        manifest["readme_file"] = str(readme_path)

        manifest_path = run_dir / "vector_export_manifest.json"
        with open(manifest_path, "w", encoding="utf-8") as fh:
            json.dump(manifest, fh, indent=2)
        manifest["manifest_file"] = str(manifest_path)

        logger.info("[VectorExport] Done. Excel workbook: %s", excel_path)
        return manifest

    @staticmethod
    def _write_binary_txt(path: Path, testset: Sequence[Sequence[int]]) -> None:
        with open(path, "w", encoding="utf-8", newline="\n") as fh:
            for vec in testset:
                fh.write(_bits_to_string(vec) + "\n")

    @staticmethod
    def _write_hex_txt(path: Path, testset: Sequence[Sequence[int]]) -> None:
        with open(path, "w", encoding="utf-8", newline="\n") as fh:
            for vec in testset:
                fh.write(_bits_to_hex(vec) + "\n")

    @staticmethod
    def _write_readme(path: Path, circuit_name: str, width: int) -> None:
        text = f"""MERS vector exports for {circuit_name}

Files in this folder:

1. *_vectors.txt
   Binary vectors, one vector per line, no header.

2. *_vectors_hex.txt
   Same vectors in hexadecimal, one vector per line, no header.

3. *_eval10k_vectors.txt and *_eval10k_vectors_hex.txt
   Only created when a full testset has more than 10,000 vectors.
   These first 10,000 vectors are the paper-style evaluation subset.

4. {circuit_name}_all_testsets.xlsx
   Compact Excel workbook with index, bitstring, hex string, and vector width.

5. primary_input_order.txt
   Mapping from bit index to primary-input / scan-input name.

Vector width: {width} bits.

IMPORTANT:
Do not shuffle the vectors.  MERS, MERS-h, and MERS-s are transition-based;
the side-channel switching is created by the exact saved order:

    zero vector -> vector 1 -> vector 2 -> vector 3 -> ...
"""
        path.write_text(text, encoding="utf-8")


# ----------------------------------------------------------------------
# Excel helpers
# ----------------------------------------------------------------------

COL_DARK_BLUE = "1F4E79"
COL_MED_BLUE = "2E75B6"
COL_LIGHT_BLUE = "D9EAF7"
COL_WHITE = "FFFFFF"
COL_GREY = "D9E1F2"


def _thin_border():
    thin = Side(style="thin", color="D0D0D0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header(cell):
    cell.fill = PatternFill(fgColor=COL_DARK_BLUE, fill_type="solid")
    cell.font = Font(color=COL_WHITE, bold=True, name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border()


def _data(cell, align="left", bold=False):
    cell.font = Font(name="Arial", size=10, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    cell.border = _thin_border()


def _setup_summary_sheet(ws):
    ws.freeze_panes = "A6"
    ws.sheet_view.showGridLines = False


def _fill_summary_sheet(ws, circuit_name: str, width: int, rows: list, readme_name: str):
    ws["A1"] = "MERS Test Vector Export Summary"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color=COL_DARK_BLUE)
    ws["A2"] = "Circuit"
    ws["B2"] = circuit_name
    ws["A3"] = "Vector width"
    ws["B3"] = width
    ws["A4"] = "README"
    ws["B4"] = readme_name

    for cell in ["A2", "A3", "A4"]:
        ws[cell].font = Font(name="Arial", size=10, bold=True)
    for cell in ["B2", "B3", "B4"]:
        ws[cell].font = Font(name="Arial", size=10)

    headers = [
        "Method", "Full vector count", "Evaluation count", "Vector width",
        "Binary TXT file", "Hex TXT file", "Excel sheet", "Separate eval10k files?"
    ]
    start_row = 6
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=col, value=h)
        _header(c)

    for r, row in enumerate(rows, start_row + 1):
        for cidx, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=cidx, value=val)
            _data(cell, align="center" if cidx in [2, 3, 4, 8] else "left")
            if r % 2 == 0:
                cell.fill = PatternFill(fgColor=COL_LIGHT_BLUE, fill_type="solid")

    widths = [14, 18, 18, 14, 34, 34, 18, 22]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.auto_filter.ref = f"A{start_row}:H{start_row}"


def _setup_input_order_sheet(ws, input_names: List[str]):
    ws.freeze_panes = "A2"
    headers = ["Bit index", "Primary input / scan input name"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        _header(c)
    for idx, name in enumerate(input_names):
        row = idx + 2
        c1 = ws.cell(row=row, column=1, value=idx)
        c2 = ws.cell(row=row, column=2, value=name)
        _data(c1, align="center")
        _data(c2)
        if row % 2 == 0:
            c1.fill = PatternFill(fgColor=COL_LIGHT_BLUE, fill_type="solid")
            c2.fill = PatternFill(fgColor=COL_LIGHT_BLUE, fill_type="solid")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 34
    ws.auto_filter.ref = "A1:B1"


def _setup_vector_sheet(ws, method: str, testset: Sequence[Sequence[int]]):
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    headers = ["Index", "Bitstring", "Hex", "Width"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        _header(c)

    for idx, vec in enumerate(testset, 1):
        bitstr = _bits_to_string(vec)
        hexstr = _bits_to_hex(vec)
        values = [idx, bitstr, hexstr, len(vec)]
        for cidx, val in enumerate(values, 1):
            cell = ws.cell(row=idx + 1, column=cidx, value=val)
            _data(cell, align="center" if cidx in [1, 4] else "left")
        if idx % 2 == 0:
            for cidx in range(1, 5):
                ws.cell(row=idx + 1, column=cidx).fill = PatternFill(fgColor=COL_LIGHT_BLUE, fill_type="solid")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 80
    ws.column_dimensions["D"].width = 12
    ws.auto_filter.ref = "A1:D1"


# ----------------------------------------------------------------------
# Generic helpers
# ----------------------------------------------------------------------

def _normalise_testset(testset: Sequence[Sequence[Any]]) -> List[List[int]]:
    out: List[List[int]] = []
    for vec in testset:
        out.append([int(x) for x in list(vec)])
    return out


def _bits_to_string(bits: Sequence[int]) -> str:
    return "".join("1" if int(b) else "0" for b in bits)


def _bits_to_hex(bits: Sequence[int]) -> str:
    bitstr = _bits_to_string(bits)
    if not bitstr:
        return ""
    pad = (-len(bitstr)) % 4
    padded = ("0" * pad) + bitstr
    return format(int(padded, 2), f"0{len(padded)//4}x")


def _safe_filename(name: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(name)).strip("_")
    return safe or "testset"


def _safe_sheet_name(name: str) -> str:
    safe = re.sub(r"[\\/*?:\[\]]+", "_", str(name)).strip()
    safe = safe or "Sheet"
    return safe[:31]
