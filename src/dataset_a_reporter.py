"""
dataset_a_reporter.py
======================
Excel report comparing Dataset A (anti-rare-cone, statistically matched)
against Dataset B (the MERS-family testset it was paired against).

Sheets
------
  * "Summary"   - aggregate (mean/std/min/max) for both datasets across
                   RareConeSwitch, BackgroundSwitch, Input HD/HW,
                   Output HD/HW, plus a bar chart of the means and a
                   "RareConeSwitch == 0" coverage row.
  * "Per-Pair"  - one row per vector pair j, A's and B's stats side by
                   side, plus the per-pair delta.
  * "Cone Nodes"- the suspicious-cone and background node lists used.
"""

import logging
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

from src.excel_reporter import (
    _hdr_style, _data_style, _thin_border, _col_width,
    COL_DARK_BLUE, COL_MED_BLUE, COL_LIGHT_BLUE, COL_GREEN, COL_RED, COL_WHITE,
)

logger = logging.getLogger(__name__)

METRICS = [
    ("rare_cone_sw", "RareConeSwitch"),
    ("bg_sw",        "BackgroundSwitch"),
    ("in_hd",        "Input Hamming Distance"),
    ("in_hw",        "Input Hamming Weight"),
    ("out_hd",       "Output Hamming Distance"),
    ("out_hw",       "Output Hamming Weight"),
]


class DatasetComparisonReporter:
    def __init__(self, output_dir: str = "mers_results"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(self, result: dict, circuit_name: str, tag: str = "") -> str:
        """
        Parameters
        ----------
        result : dict returned by DatasetAGenerator.generate(), i.e. with
                 keys per_pair_A, per_pair_B, cone_nodes, bg_nodes, weights.
        """
        per_a, per_b = result['per_pair_A'], result['per_pair_B']
        K = len(per_a)

        wb = Workbook()

        # ------------------------------------------------------------
        # Sheet 1 - Summary
        # ------------------------------------------------------------
        ws = wb.active
        ws.title = "Summary"
        ws.freeze_panes = "A2"

        headers = ["Metric", "A: mean", "A: std", "A: min", "A: max",
                   "B: mean", "B: std", "B: min", "B: max",
                   "Mean Diff (A-B)", "Mean Diff %"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            _hdr_style(c)
            c.border = _thin_border()

        means = {}
        for r, (key, label) in enumerate(METRICS, 2):
            a = np.array([d[key] for d in per_a], dtype=float)
            b = np.array([d[key] for d in per_b], dtype=float)
            means[key] = (a.mean(), b.mean())

            row_vals = [label, a.mean(), a.std(), a.min(), a.max(),
                        b.mean(), b.std(), b.min(), b.max(),
                        a.mean() - b.mean(),
                        (a.mean() - b.mean()) / b.mean() * 100 if b.mean() else 0.0]
            for col, v in enumerate(row_vals, 1):
                c = ws.cell(row=r, column=col, value=v)
                c.border = _thin_border()
                if col == 1:
                    _data_style(c, bold=True)
                elif col == 11:
                    _data_style(c, align="right", number_fmt="+0.0%")
                    c.value = row_vals[10] / 100.0
                    if key == "rare_cone_sw" and row_vals[10] < -30:
                        c.font = c.font.copy(color=COL_GREEN, bold=True)
                else:
                    _data_style(c, align="right", number_fmt="0.00")

        # Coverage row: fraction of pairs with RareConeSwitch == 0
        r = len(METRICS) + 3
        cov_a = sum(1 for d in per_a if d['rare_cone_sw'] == 0) / max(1, K)
        cov_b = sum(1 for d in per_b if d['rare_cone_sw'] == 0) / max(1, K)
        c = ws.cell(row=r, column=1, value="Pairs with RareConeSwitch == 0")
        _data_style(c, bold=True)
        c2 = ws.cell(row=r, column=2, value=cov_a)
        _data_style(c2, align="right", number_fmt="0.0%")
        c6 = ws.cell(row=r, column=6, value=cov_b)
        _data_style(c6, align="right", number_fmt="0.0%")

        r += 1
        c = ws.cell(row=r, column=1, value="Vector pairs (K)")
        _data_style(c, bold=True)
        ws.cell(row=r, column=2, value=K)

        r += 1
        c = ws.cell(row=r, column=1, value="Suspicious cone nodes")
        _data_style(c, bold=True)
        ws.cell(row=r, column=2, value=len(result['cone_nodes']))

        r += 1
        c = ws.cell(row=r, column=1, value="Background nodes")
        _data_style(c, bold=True)
        ws.cell(row=r, column=2, value=len(result['bg_nodes']))

        if result.get('weights'):
            r += 1
            c = ws.cell(row=r, column=1, value="Objective weights")
            _data_style(c, bold=True)
            ws.cell(row=r, column=2,
                    value=", ".join(f"{k}={v:g}" for k, v in result['weights'].items()))

        info = result.get('solver_info') or {}
        if info:
            r += 1
            c = ws.cell(row=r, column=1, value="Solver mode")
            _data_style(c, bold=True)
            ws.cell(row=r, column=2, value=info.get('mode', 'flat'))
            if info.get('mode') == 'phased':
                r += 1
                c = ws.cell(row=r, column=1, value="Phase 1 rounds / w_rare")
                _data_style(c, bold=True)
                ws.cell(row=r, column=2,
                        value=f"{info['phase1_rounds']} / {info['phase1_weights'].get('rare')}")
                r += 1
                c = ws.cell(row=r, column=1, value="Phase 3 rounds")
                _data_style(c, bold=True)
                ws.cell(row=r, column=2, value=info['phase3_rounds'])
                r += 1
                c = ws.cell(row=r, column=1, value="Locked cone-input PIs")
                _data_style(c, bold=True)
                ws.cell(row=r, column=2, value=info['n_cone_input_pis_locked'])
                r += 1
                c = ws.cell(row=r, column=1, value="Free background PIs (Phase 3)")
                _data_style(c, bold=True)
                ws.cell(row=r, column=2, value=info['n_background_pis_free'])

        _col_width(ws, 1, 28)
        for col in range(2, 12):
            _col_width(ws, col, 13)

        # Bar chart of means (A vs B) for the 6 switching/HD/HW metrics
        chart = BarChart()
        chart.title = "Dataset A vs Dataset B - mean per-pair metrics"
        chart.y_axis.title = "Mean count per vector pair"
        chart.style = 10
        data = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=1 + len(METRICS))
        data_b = Reference(ws, min_col=6, max_col=6, min_row=1, max_row=1 + len(METRICS))
        cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(METRICS))
        chart.add_data(data, titles_from_data=True)
        chart.add_data(data_b, titles_from_data=True)
        chart.set_categories(cats)
        chart.width, chart.height = 18, 10
        ws.add_chart(chart, "A" + str(len(METRICS) + 8))

        # ------------------------------------------------------------
        # Sheet 2 - Per-Pair
        # ------------------------------------------------------------
        ws2 = wb.create_sheet("Per-Pair")
        ws2.freeze_panes = "B2"

        pp_headers = ["Pair #"]
        for _key, label in METRICS:
            pp_headers += [f"A {label}", f"B {label}"]
        for col, h in enumerate(pp_headers, 1):
            c = ws2.cell(row=1, column=col, value=h)
            _hdr_style(c, bg=COL_MED_BLUE)
            c.border = _thin_border()

        for j in range(K):
            row = j + 2
            ws2.cell(row=row, column=1, value=j + 1).border = _thin_border()
            col = 2
            for key, _label in METRICS:
                ca = ws2.cell(row=row, column=col, value=per_a[j][key])
                cb = ws2.cell(row=row, column=col + 1, value=per_b[j][key])
                ca.border = _thin_border()
                cb.border = _thin_border()
                if key == "rare_cone_sw":
                    if per_a[j][key] == 0:
                        ca.font = ca.font.copy(color=COL_GREEN, bold=True)
                col += 2

        _col_width(ws2, 1, 8)
        for col in range(2, len(pp_headers) + 1):
            _col_width(ws2, col, 12)

        # ------------------------------------------------------------
        # Sheet 3 - Cone Nodes
        # ------------------------------------------------------------
        ws3 = wb.create_sheet("Cone Nodes")
        for col, h in enumerate(["Suspicious Cone Node", "Background Node"], 1):
            c = ws3.cell(row=1, column=col, value=h)
            _hdr_style(c)
            c.border = _thin_border()
        cone_nodes, bg_nodes = result['cone_nodes'], result['bg_nodes']
        for r in range(max(len(cone_nodes), len(bg_nodes))):
            if r < len(cone_nodes):
                ws3.cell(row=r + 2, column=1, value=cone_nodes[r]).border = _thin_border()
            if r < len(bg_nodes):
                ws3.cell(row=r + 2, column=2, value=bg_nodes[r]).border = _thin_border()
        _col_width(ws3, 1, 28)
        _col_width(ws3, 2, 28)

        safe = circuit_name.replace(" ", "_")
        suffix = f"_{tag}" if tag else ""
        path = self.output_dir / f"{safe}_dataset_A_vs_B{suffix}.xlsx"
        wb.save(path)
        logger.info(f"[Reporter] Saved: {path}")
        return str(path)
