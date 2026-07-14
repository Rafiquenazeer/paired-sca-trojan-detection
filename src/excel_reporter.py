"""
excel_reporter.py
=================
Generates formatted Excel workbooks:

  1. <circuit>_rare_nodes.xlsx
       • Sheet "Rare Nodes"        – all rare nodes with probability, colour-coded
       • Sheet "All Node Probs"    – P(node=1) for every internal node
       • Sheet "Summary"           – circuit statistics
       • Sheet "Probability Chart" – bar chart of rare-node probabilities

  2. <circuit>_<N>trig_sensitivity.xlsx
       • Sheet "Side Channel Sensitivity" – SCS comparison across methods
       • Sheet "Delta Switch"             – DeltaSwitch comparison
       • Sheet "Improvement %"           – % improvement vs Random / MERO
"""

import logging
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

logger = logging.getLogger(__name__)

# -----------------------------------------------------------------------
# Colour palette
# -----------------------------------------------------------------------
COL_DARK_BLUE  = "1F4E79"   # header background
COL_MED_BLUE   = "2E75B6"   # sub-header
COL_LIGHT_BLUE = "BDD7EE"   # alternate row tint
COL_RED        = "FF4444"    # very rare (prob ≤ 0.05)
COL_ORANGE     = "FF8800"    # quite rare (0.05 < prob ≤ 0.10)
COL_YELLOW     = "FFD966"    # note / highlight
COL_GREEN      = "00B050"    # improvement positive
COL_WHITE      = "FFFFFF"


def _hdr_style(cell, bg=COL_DARK_BLUE, fg=COL_WHITE, bold=True, wrap=True):
    cell.fill = PatternFill(fgColor=bg, fill_type="solid")
    cell.font = Font(color=fg, bold=bold, name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _data_style(cell, bold=False, align="left", number_fmt=None):
    cell.font = Font(name="Arial", size=10, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if number_fmt:
        cell.number_format = number_fmt


def _thin_border():
    thin = Side(style="thin", color="D0D0D0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ======================================================================
# Report 1 – Rare Nodes
# ======================================================================

class RareNodeReporter:
    def __init__(self, output_dir: str = "mers_results"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(self, rare_nodes_info: dict, circuit, circuit_name: str) -> str:
        """Create and save the rare-node Excel workbook. Returns file path."""
        rare_nodes  = rare_nodes_info['rare_nodes']
        node_probs  = rare_nodes_info['node_probs']

        wb = Workbook()

        # ----------------------------------------------------------------
        # Sheet 1 – Rare Nodes
        # ----------------------------------------------------------------
        ws1 = wb.active
        ws1.title = "Rare Nodes"
        ws1.freeze_panes = "A2"

        headers = [
            "Node Name", "P(rare value)", "Rare Value", "Non-Rare Value",
            "Category", "Is Primary Output", "Rank (rarest first)"
        ]
        for col, h in enumerate(headers, 1):
            c = ws1.cell(row=1, column=col, value=h)
            _hdr_style(c)
            c.border = _thin_border()

        sorted_rare = sorted(rare_nodes.items(), key=lambda x: x[1]['probability'])
        po_set = set(circuit.primary_outputs)

        for rank, (node, info) in enumerate(sorted_rare, 1):
            row = rank + 1
            prob = info['probability']
            rv   = info['rare_value']
            cat  = f"Rare-{rv} (P({rv}) ≤ {rare_nodes_info['rare_threshold']:.0%})"

            data = [node, prob, rv, info['non_rare_value'], cat,
                    node in po_set, rank]

            for col, val in enumerate(data, 1):
                c = ws1.cell(row=row, column=col, value=val)
                c.border = _thin_border()
                if col == 1:
                    _data_style(c, bold=True)
                elif col == 2:
                    _data_style(c, align="right", number_fmt="0.000000")
                else:
                    _data_style(c, align="center")

            # Colour-code probability cell
            if prob <= 0.05:
                bg = COL_RED
            elif prob <= 0.10:
                bg = COL_ORANGE
            else:
                bg = COL_YELLOW
            ws1.cell(row=row, column=2).fill = PatternFill(fgColor=bg, fill_type="solid")

            # Alternate row background
            if rank % 2 == 0:
                for col in [1, 3, 4, 5, 6, 7]:
                    ws1.cell(row=row, column=col).fill = \
                        PatternFill(fgColor=COL_LIGHT_BLUE, fill_type="solid")

        col_widths = [22, 18, 14, 18, 35, 18, 22]
        for i, w in enumerate(col_widths, 1):
            _col_width(ws1, i, w)

        ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        # ----------------------------------------------------------------
        # Sheet 2 – All Node Probabilities
        # ----------------------------------------------------------------
        ws2 = wb.create_sheet("All Node Probs")
        ws2.freeze_panes = "A2"

        for col, h in enumerate(["Node", "P(node=1)", "Is Rare", "Candidate Status"], 1):
            c = ws2.cell(row=1, column=col, value=h)
            _hdr_style(c, bg=COL_MED_BLUE)

        rare_set = set(rare_nodes.keys())
        stuck_set = set(rare_nodes_info.get('stuck_nodes', {}).keys())
        artifact_set = set(rare_nodes_info.get('artifact_nodes', {}).keys())
        for row, (node, prob) in enumerate(
                sorted(node_probs.items(), key=lambda x: x[1]), 2):
            is_rare = node in rare_set
            if node in artifact_set:
                status = "EXCLUDED: synthetic artefact"
            elif node in stuck_set:
                status = "EXCLUDED: stuck/constant"
            elif is_rare:
                status = "RARE TARGET"
            else:
                status = "normal"
            ws2.cell(row=row, column=1, value=node)
            c2 = ws2.cell(row=row, column=2, value=prob)
            c2.number_format = "0.000000"
            ws2.cell(row=row, column=3, value="YES" if is_rare else "")
            ws2.cell(row=row, column=4, value=status)
            if is_rare:
                for col in (1, 2, 3, 4):
                    ws2.cell(row=row, column=col).fill = \
                        PatternFill(fgColor=COL_YELLOW, fill_type="solid")

        _col_width(ws2, 1, 28)
        _col_width(ws2, 2, 16)
        _col_width(ws2, 3, 10)
        _col_width(ws2, 4, 28)
        ws2.auto_filter.ref = "A1:D1"

        # ----------------------------------------------------------------
        # Sheet 3 – Summary
        # ----------------------------------------------------------------
        ws3 = wb.create_sheet("Summary")

        r1c = sum(1 for v in rare_nodes.values() if v['rare_value'] == 1)
        r0c = len(rare_nodes) - r1c

        stuck_nodes = rare_nodes_info.get('stuck_nodes', {})
        artifact_nodes = rare_nodes_info.get('artifact_nodes', {})
        artifact_prefixes = rare_nodes_info.get('artifact_prefixes', [])
        rows = [
            ("Benchmark", circuit_name),
            ("Gate Count", len(circuit.gates)),
            ("Primary Inputs", len(circuit.primary_inputs)),
            ("Primary Outputs", len(circuit.primary_outputs)),
            ("Unresolved Gates (P=0 artefacts)", len(getattr(circuit, 'unresolved_gates', []))),
            ("Suspected Stuck Nodes (excluded)", len(stuck_nodes)),
            ("Synthetic Artefact Nodes (excluded)", len(artifact_nodes)),
            ("Artefact Prefix Filter", ", ".join(artifact_prefixes) if artifact_prefixes else "Disabled"),
            ("Total Rare Nodes (after exclusions)", len(rare_nodes)),
            ("  Rare-1 (signal=1 is rare)", r1c),
            ("  Rare-0 (signal=0 is rare)", r0c),
            ("Rare Threshold", f"{rare_nodes_info['rare_threshold']:.0%}"),
            ("Monte-Carlo Vectors", rare_nodes_info['num_vectors']),
            ("Rarest Node (prob)", sorted_rare[0][0] if sorted_rare else "N/A"),
            ("Rarest Probability", sorted_rare[0][1]['probability']
             if sorted_rare else "N/A"),
        ]

        _hdr_style(ws3.cell(row=1, column=1, value="Metric"), bg=COL_DARK_BLUE)
        _hdr_style(ws3.cell(row=1, column=2, value="Value"), bg=COL_DARK_BLUE)

        for i, (metric, value) in enumerate(rows, 2):
            ws3.cell(row=i, column=1, value=metric).font = Font(name="Arial", size=10)
            ws3.cell(row=i, column=2, value=value).font = Font(name="Arial", size=10, bold=True)

        _col_width(ws3, 1, 35)
        _col_width(ws3, 2, 20)

        # ----------------------------------------------------------------
        # Sheet 4 – Chart
        # ----------------------------------------------------------------
        ws4 = wb.create_sheet("Probability Chart")
        ws4["A1"] = "Node"
        ws4["B1"] = "P(rare value)"

        for i, (node, info) in enumerate(sorted_rare[:50], 2):  # top 50 rarest
            ws4.cell(row=i, column=1, value=node)
            ws4.cell(row=i, column=2, value=info['probability'])

        n_chart = min(50, len(sorted_rare))
        chart = BarChart()
        chart.type = "col"
        chart.title = f"Top-{n_chart} Rarest Nodes – {circuit_name}"
        chart.y_axis.title = "P(rare value)"
        chart.x_axis.title = "Node"
        chart.height = 14
        chart.width  = 24
        chart.style  = 10
        chart.grouping = "clustered"
        chart.overlap  = 100

        data_ref = Reference(ws4, min_col=2, max_col=2,
                             min_row=1, max_row=n_chart + 1)
        cats_ref = Reference(ws4, min_col=1, min_row=2, max_row=n_chart + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws4.add_chart(chart, "D2")

        # ----------------------------------------------------------------
        # Save
        # ----------------------------------------------------------------
        out = self.output_dir / f"{circuit_name}_rare_nodes.xlsx"
        wb.save(str(out))
        logger.info(f"[Reporter] Saved: {out}")
        return str(out)


# ======================================================================
# Report 2 – Sensitivity Comparison
# ======================================================================

class SensitivityReporter:
    def __init__(self, output_dir: str = "mers_results"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(self, results: dict, circuit_name: str,
                 num_triggers: int) -> str:
        """
        *results* structure:
        {
          'Random': {avg_max_delta, avg_avg_delta, avg_max_relative, avg_avg_relative, testset_size},
          'MERS':   {...},
          'MERS-h': {...},
          'MERS-s': {...},
        }
        """
        wb = Workbook()
        methods = list(results.keys())

        # ----------------------------------------------------------------
        # Sheet 1 – Side Channel Sensitivity
        # ----------------------------------------------------------------
        ws1 = wb.active
        ws1.title = "Side Channel Sensitivity"
        ws1.freeze_panes = "B2"

        headers = ["Method", "Avg MaxRelativeSwitch (SCS)", "Avg AvgRelativeSwitch",
                   "Testset Size", "vs Random (%)", "vs MERO (%)"]

        for col, h in enumerate(headers, 1):
            c = ws1.cell(row=1, column=col, value=h)
            _hdr_style(c)

        rand_scs = results.get('Random', {}).get('avg_max_relative', 1e-9)
        mero_scs = results.get('MERO', {}).get('avg_max_relative', None)

        for row, method in enumerate(methods, 2):
            m = results[method]
            scs = m.get('avg_max_relative', 0)
            avg_rel = m.get('avg_avg_relative', 0)
            ts  = m.get('testset_size', 0)
            # BUG FIX: Excel "+0.00%" format multiplies by 100 automatically;
            # write the raw ratio (0.0678), not the %-scaled value (6.78)
            vs_rand = (scs - rand_scs) / max(rand_scs, 1e-9)

            vs_mero = "–" if mero_scs is None else (scs - mero_scs) / max(mero_scs, 1e-9)
            data = [method, scs, avg_rel, ts, vs_rand, vs_mero]
            fmts = [None, "0.00000", "0.00000", "#,##0", "+0.00%;-0.00%",
                    None if isinstance(vs_mero, str) else "+0.00%;-0.00%"]

            for col, (val, fmt) in enumerate(zip(data, fmts), 1):
                c = ws1.cell(row=row, column=col, value=val)
                _data_style(c, align="center")
                if fmt:
                    c.number_format = fmt

            # Green background for best SCS
            if method != 'Random' and scs > rand_scs:
                ws1.cell(row=row, column=2).fill = \
                    PatternFill(fgColor=COL_GREEN, fill_type="solid")
                ws1.cell(row=row, column=2).font = \
                    Font(color=COL_WHITE, bold=True, name="Arial", size=10)

        for i, w in enumerate([14, 28, 24, 14, 16, 14], 1):
            _col_width(ws1, i, w)

        # ----------------------------------------------------------------
        # Sheet 2 – Delta Switch
        # ----------------------------------------------------------------
        ws2 = wb.create_sheet("Delta Switch")

        headers2 = ["Method", "Avg MaxDeltaSwitch", "Avg AvgDeltaSwitch", "Testset Size"]
        for col, h in enumerate(headers2, 1):
            c = ws2.cell(row=1, column=col, value=h)
            _hdr_style(c, bg=COL_MED_BLUE)

        for row, method in enumerate(methods, 2):
            m = results[method]
            data = [method,
                    m.get('avg_max_delta', 0),
                    m.get('avg_avg_delta', 0),
                    m.get('testset_size', 0)]
            fmts = [None, "0.000", "0.000", "#,##0"]
            for col, (val, fmt) in enumerate(zip(data, fmts), 1):
                c = ws2.cell(row=row, column=col, value=val)
                _data_style(c, align="center")
                if fmt:
                    c.number_format = fmt

        for i, w in enumerate([14, 22, 22, 14], 1):
            _col_width(ws2, i, w)

        # ----------------------------------------------------------------
        # Chart sheet
        # ----------------------------------------------------------------
        ws3 = wb.create_sheet("Chart")
        ws3["A1"] = "Method"
        ws3["B1"] = "SCS (Avg MaxRelativeSwitch)"
        for i, method in enumerate(methods, 2):
            ws3.cell(row=i, column=1, value=method)
            ws3.cell(row=i, column=2,
                     value=results[method].get('avg_max_relative', 0))

        chart = BarChart()
        chart.title = (f"Side-Channel Sensitivity – {circuit_name} "
                       f"({num_triggers}-trigger Trojans)")
        chart.y_axis.title = "SCS (MaxRelativeSwitch)"
        chart.x_axis.title = "Test Method"
        chart.height = 12
        chart.width  = 20
        chart.style  = 10

        data_r = Reference(ws3, min_col=2, max_col=2,
                           min_row=1, max_row=len(methods) + 1)
        cats_r = Reference(ws3, min_col=1, min_row=2, max_row=len(methods) + 1)
        chart.add_data(data_r, titles_from_data=True)
        chart.set_categories(cats_r)
        ws3.add_chart(chart, "D2")

        # ----------------------------------------------------------------
        # Save
        # ----------------------------------------------------------------
        out = self.output_dir / f"{circuit_name}_{num_triggers}trig_sensitivity.xlsx"
        wb.save(str(out))
        logger.info(f"[Reporter] Saved: {out}")
        return str(out)


# ======================================================================
# Report 3 – Rare-node Regions
# ======================================================================

class RegionReporter:
    """
    Workbook describing the rare-node region decomposition:

      • Sheet "Summary"        – run parameters + per-region overview table
                                 with a bar chart of region sizes, and a
                                 top-K coverage table.
      • Sheet "All Rare Nodes" – every rare node, its probability/rare-value,
                                 and which region it was assigned to.
      • Sheet "Region <id>"    – one sheet per region (top regions), listing
                                 that region's rare nodes.
    """

    def __init__(self, output_dir: str = "mers_results"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(self, region_result: dict, rare_nodes: dict,
                 circuit_name: str, per_region_sheets: int = 10,
                 top_k_for_coverage=(1, 2, 3, 5, 10)) -> str:
        regions    = region_result["regions"]
        singletons = region_result.get("singletons", [])
        params     = region_result.get("params", {})
        total_rare = len(rare_nodes)

        # Map each rare node -> region id (or -1 for singleton/unassigned).
        node_region = {}
        for reg in regions:
            for n in reg["rare_nodes"]:
                node_region[n] = reg["id"]
        for n in singletons:
            node_region.setdefault(n, -1)

        wb = Workbook()

        # ---------------- Sheet 1: Summary ----------------
        ws = wb.active
        ws.title = "Summary"
        r = 1
        c = ws.cell(row=r, column=1, value=f"Rare-node Regions — {circuit_name}")
        _hdr_style(c, bg=COL_DARK_BLUE)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 2

        stats = [
            ("Total rare nodes", total_rare),
            ("Total regions", len(regions)),
            ("Singleton rare nodes", len(singletons)),
            ("Largest region (rare nodes)", regions[0]["size"] if regions else 0),
            ("Region mode", params.get("mode", "")),
        ]
        if params.get("kind") == "physical":
            stats += [
                ("Region kind", "physical (FPGA placement)"),
                ("Clustering radius", params.get("radius", "")),
                ("Max region size", params.get("max_region_size", "")),
            ]
        else:
            stats += [
                ("Region kind", "logical (fan-in cone)"),
                ("Fan-in depth", params.get("fanin_depth", "")),
                ("Overlap threshold", params.get("overlap_threshold", "")),
                ("Max region size", params.get("max_region_size", "")),
            ]
        for label, val in stats:
            cc = ws.cell(row=r, column=1, value=label); _data_style(cc, bold=True)
            cc.border = _thin_border()
            cv = ws.cell(row=r, column=2, value=val); _data_style(cv)
            cv.border = _thin_border()
            r += 1
        r += 1

        # Coverage table (top-K -> how many rare nodes / % / reduction)
        cov_hdr_row = r
        for i, h in enumerate(["Top-K regions", "Rare nodes covered",
                               "% of all rare", "MERS reduction"]):
            cc = ws.cell(row=r, column=1 + i, value=h); _hdr_style(cc, bg=COL_MED_BLUE)
            cc.border = _thin_border()
        r += 1
        for k in top_k_for_coverage:
            if k > len(regions):
                continue
            covered = sum(len(regions[i]["rare_nodes"]) for i in range(k))
            pct = covered / total_rare if total_rare else 0
            cc = ws.cell(row=r, column=1, value=f"top-{k}"); _data_style(cc); cc.border = _thin_border()
            cc = ws.cell(row=r, column=2, value=covered); _data_style(cc, align="center"); cc.border = _thin_border()
            cc = ws.cell(row=r, column=3, value=pct); _data_style(cc, align="center", number_fmt="0.0%"); cc.border = _thin_border()
            cc = ws.cell(row=r, column=4, value=(1 - pct)); _data_style(cc, align="center", number_fmt="0.0%"); cc.border = _thin_border()
            r += 1
        r += 1

        # Per-region overview table
        ov_hdr_row = r
        is_phys = params.get("kind") == "physical"
        size_col = "Slices (footprint)" if is_phys else "Cone gates"
        dens_col = "Density (rare/slice)" if is_phys else "Density (rare/cone)"
        headers = ["Region", "Rare nodes", size_col, dens_col,
                   "Sample rare nodes"]
        for i, h in enumerate(headers):
            cc = ws.cell(row=r, column=1 + i, value=h); _hdr_style(cc)
            cc.border = _thin_border()
        r += 1
        first_data_row = r
        for reg in regions:
            cc = ws.cell(row=r, column=1, value=reg["id"]); _data_style(cc, align="center"); cc.border = _thin_border()
            cc = ws.cell(row=r, column=2, value=reg["size"]); _data_style(cc, align="center"); cc.border = _thin_border()
            cc = ws.cell(row=r, column=3, value=reg["cone_size"]); _data_style(cc, align="center"); cc.border = _thin_border()
            cc = ws.cell(row=r, column=4, value=round(reg["density"], 4)); _data_style(cc, align="center", number_fmt="0.0000"); cc.border = _thin_border()
            cc = ws.cell(row=r, column=5, value=", ".join(reg["rare_nodes"][:4])); _data_style(cc); cc.border = _thin_border()
            # tint the densest region rows lightly
            if reg["id"] < 3:
                for col in range(1, 6):
                    ws.cell(row=r, column=col).fill = PatternFill(
                        fgColor=COL_LIGHT_BLUE, fill_type="solid")
            r += 1
        last_data_row = r - 1

        for col, w in [(1, 10), (2, 12), (3, 12), (4, 18), (5, 50)]:
            _col_width(ws, col, w)

        # Bar chart of region sizes (rare nodes per region)
        if regions:
            chart = BarChart()
            chart.title = "Rare nodes per region (densest first)"
            chart.y_axis.title = "Rare nodes"
            chart.x_axis.title = "Region id"
            data = Reference(ws, min_col=2, min_row=ov_hdr_row,
                             max_row=last_data_row)
            cats = Reference(ws, min_col=1, min_row=first_data_row,
                             max_row=last_data_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 8
            chart.width = 18
            chart.legend = None
            ws.add_chart(chart, f"G{cov_hdr_row}")

        # ---------------- Sheet 2: All Rare Nodes ----------------
        ws2 = wb.create_sheet("All Rare Nodes")
        hdrs = ["Rare node", "P(rare value)", "Rare value", "Region id",
                "In top region?"]
        for i, h in enumerate(hdrs):
            cc = ws2.cell(row=1, column=1 + i, value=h); _hdr_style(cc)
            cc.border = _thin_border()
        rr = 2
        # sort: by region id (densest first), then name; singletons last
        def sort_key(item):
            n = item[0]
            reg = node_region.get(n, -1)
            return (reg if reg >= 0 else 10**9, n)
        for n, info in sorted(rare_nodes.items(), key=sort_key):
            reg_id = node_region.get(n, -1)
            prob = info.get("probability", "")
            rv = info.get("rare_value", "")
            cc = ws2.cell(row=rr, column=1, value=n); _data_style(cc); cc.border = _thin_border()
            cc = ws2.cell(row=rr, column=2, value=prob); _data_style(cc, align="center",
                number_fmt="0.000000" if isinstance(prob, float) else None); cc.border = _thin_border()
            cc = ws2.cell(row=rr, column=3, value=rv); _data_style(cc, align="center"); cc.border = _thin_border()
            cc = ws2.cell(row=rr, column=4, value=(reg_id if reg_id >= 0 else "singleton"))
            _data_style(cc, align="center"); cc.border = _thin_border()
            cc = ws2.cell(row=rr, column=5, value=("yes" if 0 <= reg_id < 3 else ""))
            _data_style(cc, align="center"); cc.border = _thin_border()
            rr += 1
        for col, w in [(1, 34), (2, 14), (3, 11), (4, 11), (5, 14)]:
            _col_width(ws2, col, w)

        # ---------------- Sheet(s) 3..: per-region ----------------
        for reg in regions[:per_region_sheets]:
            ws3 = wb.create_sheet(f"Region {reg['id']}")
            t = ws3.cell(row=1, column=1,
                         value=f"Region {reg['id']} — {reg['size']} rare nodes, "
                               f"{reg['cone_size']} cone gates, "
                               f"density {reg['density']:.4f}")
            _hdr_style(t, bg=COL_MED_BLUE)
            ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
            for i, h in enumerate(["#", "Rare node", "P(rare value)"]):
                cc = ws3.cell(row=2, column=1 + i, value=h); _hdr_style(cc)
                cc.border = _thin_border()
            rr = 3
            for k, n in enumerate(reg["rare_nodes"], 1):
                prob = rare_nodes.get(n, {}).get("probability", "")
                cc = ws3.cell(row=rr, column=1, value=k); _data_style(cc, align="center"); cc.border = _thin_border()
                cc = ws3.cell(row=rr, column=2, value=n); _data_style(cc); cc.border = _thin_border()
                cc = ws3.cell(row=rr, column=3, value=prob); _data_style(cc, align="center",
                    number_fmt="0.000000" if isinstance(prob, float) else None); cc.border = _thin_border()
                rr += 1
            for col, w in [(1, 6), (2, 34), (3, 14)]:
                _col_width(ws3, col, w)

        out = self.output_dir / f"{circuit_name}_rare_node_regions.xlsx"
        wb.save(str(out))
        logger.info(f"[Reporter] Saved: {out}")
        return str(out)
